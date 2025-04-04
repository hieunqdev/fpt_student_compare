from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, BackgroundTasks
from app.models.poly_cong_nhan_sinh_vien import PolyCongNhanSinhVien
from app.utils.pdf_reader import poly_cong_nhan_sinh_vien_lay_danh_sach_sinh_vien_tu_pdf
from pydantic import BaseModel
from app.database import SessionLocal
from sqlmodel import Session
import pandas as pd
import openpyxl
from io import BytesIO
import time
from fastapi.responses import StreamingResponse

router = APIRouter()

class GoogleSheetRequest(BaseModel):
    sheet_url: str

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def normalize_text(text):
    """Chuẩn hóa văn bản: loại bỏ khoảng trắng hai đầu, xuống dòng và viết thường"""
    return text.strip().replace("\n", " ").lower() if text else ""

# Poly compare_students
def poly_compare_students(ds_sinh_vien_sheets, ds_sinh_vien_pdf):
    """
    So sánh danh sách sinh viên từ Excel và từ database.
    """
    cap_nhat_ket_qua = []
    try:
        # Duyệt qua danh sách sinh viên từ Excel
        for index, sv_sheets in ds_sinh_vien_sheets.iterrows():
            mssv_excel = normalize_text(str(sv_sheets['MSSV']).strip())

            for sv in ds_sinh_vien_pdf:
                if normalize_text(str(sv.mssv).strip()) == mssv_excel:
                    sv_pdf = sv
                    break

            # Nếu không tìm thấy sinh viên trong database, bỏ qua
            if sv_pdf is None:
                continue

            khac_biet = []

            # So sánh từng thuộc tính của sinh viên
            if normalize_text(sv_sheets['Họ Tên SV']) != normalize_text(sv_pdf.ho_va_ten):
                khac_biet.append(f"({sv_sheets['Họ Tên SV']} != {sv_pdf.ho_va_ten})")

            # if normalize_text(str(sv_sheets['Ngày sinh'])) != normalize_text(sv_pdf.ngay_sinh):
            #     khac_biet.append(f"({str(sv_sheets['Ngày sinh'])} != {sv_pdf.ngay_sinh})")
            #
            # if normalize_text(str(sv_sheets['Giới tính'])) != normalize_text(sv_pdf.gioi_tinh):
            #     khac_biet.append(f"({str(sv_sheets['Giới tính'])} != {sv_pdf.gioi_tinh})")
            #
            # if normalize_text(str(sv_sheets['Dân tộc'])) != normalize_text(sv_pdf.dan_toc):
            #     khac_biet.append(f"({str(sv_sheets['Dân tộc'])} != {sv_pdf.dan_toc})")

            if khac_biet:
                cap_nhat_ket_qua.append({
                    "mssv": mssv_excel,
                    "ngay_sinh": sv_pdf.ngay_sinh,
                    "gioi_tinh": sv_pdf.gioi_tinh,
                    "dan_toc": sv_pdf.dan_toc,
                    "ghi_chu": "; ".join(khac_biet)
                })

        return cap_nhat_ket_qua
    except Exception as e:
        raise Exception(f"Lỗi khi so sánh sinh viên: {str(e)}")

def cap_nhat_khac_biet_vao_excel(file_content: bytes, cap_nhat_ket_qua: list, cot_quyet_dinh: str, cot_ghi_chu: str,
                                 so_qd: str, ngay_sinh: str, gioi_tinh: str, dan_toc: str):
    """
    Cập nhật thông tin khác biệt vào file Excel.
    """
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        sheet = workbook.active

        # Tìm chỉ số cột từ ký tự cột (A -> 1, B -> 2, ...)
        col_number = openpyxl.utils.column_index_from_string(cot_ghi_chu)
        col_ngay_sinh = openpyxl.utils.column_index_from_string(ngay_sinh)
        col_gioi_tinh = openpyxl.utils.column_index_from_string(gioi_tinh)
        col_dan_toc = openpyxl.utils.column_index_from_string(dan_toc)
        col_number_cot_quyet_dinh = openpyxl.utils.column_index_from_string(cot_quyet_dinh)

        # Duyệt qua từng sinh viên và cập nhật khác biệt vào Excel
        for cap_nhat in cap_nhat_ket_qua:
            mssv = cap_nhat.get("mssv")
            ngay_sinh = cap_nhat.get("ngay_sinh")
            gioi_tinh = cap_nhat.get("gioi_tinh")
            dan_toc = cap_nhat.get("dan_toc")
            ghi_chu = cap_nhat.get("ghi_chu")

            # Duyệt qua từng hàng để tìm MSSV
            for row in range(2, sheet.max_row + 1):  # Bỏ qua hàng đầu tiên (tiêu đề)
                if normalize_text(str(sheet.cell(row=row, column=2).value)) == str(mssv):
                    sheet.cell(row=row, column=col_number_cot_quyet_dinh, value=str(so_qd))
                    sheet.cell(row=row, column=col_ngay_sinh, value=str(ngay_sinh))
                    sheet.cell(row=row, column=col_gioi_tinh, value=str(gioi_tinh))
                    sheet.cell(row=row, column=col_dan_toc, value=str(dan_toc))
                    sheet.cell(row=row, column=col_number, value=ghi_chu)
                    break

        # Lưu file Excel sau khi cập nhật
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    except Exception as e:
        raise Exception(f"Lỗi khi cập nhật Excel: {str(e)}")

############################################### Poly
# Poly Quyết định Công nhận sinh viên
# API Poly Công nhận sinh viên
# @router.post("/upload/sheet-poly-cong-nhan-sinh-vien")
# async def poly_cong_nhan_sinh_vien_compare(file: UploadFile, cot_quyet_dinh: str = "R", cot_ghi_chu: str = "AB",
#                                            db: Session = Depends(get_db), files: list[UploadFile] = File(...),
#                                            background_tasks: BackgroundTasks = BackgroundTasks()):
#     """
#         API
#         Tải lên file PDF Quyết định công nhận sinh viên và lưu vào PostgreSQL.
#
#         Nhập danh sách sinh viên từ Excel, so sánh với danh sách từ database,
#         và cập nhật thông tin khác biệt vào cột chỉ định trong Excel.
#     """
#     record = db.query(PolyCongNhanSinhVien).filter(PolyCongNhanSinhVien.ten_file == files[0].filename).first()
#     if not record:
#         file_data_list = [(await file.read(), file.filename) for file in files]  # Đọc file ngay tại đây
#
#         # def process_and_save_pdfs():
#         danh_sach_tat_ca_sinh_vien = []
#         for file_bytes, filename in file_data_list:
#             danh_sach_sinh_vien = poly_cong_nhan_sinh_vien_lay_danh_sach_sinh_vien_tu_pdf(file_bytes, filename, db)
#             danh_sach_tat_ca_sinh_vien.extend(danh_sach_sinh_vien)
#
#         db.bulk_save_objects(danh_sach_tat_ca_sinh_vien)
#         db.commit()
#
#         # Chạy xử lý dữ liệu trong nền
#         # background_tasks.add_task(process_and_save_pdfs)
#     try:
#         # 1. Đọc danh sách sinh viên từ file Excel
#         content = await file.read()
#         ds_sinh_vien_sheets = pd.read_excel(BytesIO(content), dtype=str)
#
#         # 2. Lấy danh sách sinh viên từ database (PDF đã upload trước đó)
#         # ds_sinh_vien_pdf = db.query(SinhVien).all()
#         ds_sinh_vien_pdf = (db.query(PolyCongNhanSinhVien)
#                             .filter(PolyCongNhanSinhVien.ten_file == files[0].filename)
#                             .all())
#         so_qd = ds_sinh_vien_pdf[0].so_qd
#         # 3. So sánh danh sách sinh viên
#         cap_nhat_ket_qua = poly_compare_students(ds_sinh_vien_sheets, ds_sinh_vien_pdf)
#
#         # 4. Cập nhật khác biệt vào file Excel đã tải về
#         updated_excel = cap_nhat_khac_biet_vao_excel(content, cap_nhat_ket_qua, cot_quyet_dinh, cot_ghi_chu, so_qd)
#
#         # Trả file Excel đã cập nhật về dưới dạng StreamingResponse
#         return StreamingResponse(
#             BytesIO(updated_excel),
#             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             headers={"Content-Disposition": f"attachment; filename=updated_sinh_vien.xlsx"}
#         )
#
#     except Exception as e:
#         raise HTTPException(status_code=400, detail=str(e))


@router.post("/upload/poly-cong-nhan-sinh-vien")
async def poly_cong_nhan_sinh_vien(file: UploadFile, cot_quyet_dinh: str = "R", cot_ghi_chu: str = "AB",
                                   ngay_sinh: str = "O", gioi_tinh: str = "P", dan_toc: str = "Q",
                                   db: Session = Depends(get_db), files: list[UploadFile] = File(...)):
    record = db.query(PolyCongNhanSinhVien).filter(PolyCongNhanSinhVien.ten_file == files[0].filename).first()
    if not record:
        file_data_list = [(await file.read(), file.filename) for file in files]

        danh_sach_tat_ca_sinh_vien = []
        for file_bytes, filename in file_data_list:
            danh_sach_sinh_vien = poly_cong_nhan_sinh_vien_lay_danh_sach_sinh_vien_tu_pdf(file_bytes, filename, db)
            danh_sach_tat_ca_sinh_vien.extend(danh_sach_sinh_vien)

        db.bulk_save_objects(danh_sach_tat_ca_sinh_vien)
        db.commit()

    try:
        # 1. Đọc danh sách sinh viên từ file Excel
        content = await file.read()
        ds_sinh_vien_sheets = pd.read_excel(BytesIO(content), dtype=str)

        # 2. Lấy danh sách sinh viên từ database (PDF đã upload trước đó)
        ds_sinh_vien_pdf = (db.query(PolyCongNhanSinhVien)
                            .filter(PolyCongNhanSinhVien.ten_file == files[0].filename)
                            .all())
        so_qd = ds_sinh_vien_pdf[0].so_qd
        # 3. So sánh danh sách sinh viên
        cap_nhat_ket_qua = poly_compare_students(ds_sinh_vien_sheets, ds_sinh_vien_pdf)

        # 4. Cập nhật khác biệt vào file Excel đã tải về
        updated_excel = cap_nhat_khac_biet_vao_excel(content, cap_nhat_ket_qua, cot_quyet_dinh, cot_ghi_chu, so_qd,
                                                     ngay_sinh, gioi_tinh, dan_toc)

        # Trả file Excel đã cập nhật về dưới dạng StreamingResponse
        return StreamingResponse(
            BytesIO(updated_excel),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=updated_sinh_vien.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
